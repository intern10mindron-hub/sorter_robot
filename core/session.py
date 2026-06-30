from __future__ import annotations
import os, time
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import EXPORTS_DIR
from hardware.robot_controller import SLOT_ORDER


@dataclass
class DiamondRecord:
    diamond_id:   str
    slot_index:   int
    weight_ct:    float
    timestamp:    str
    tray_x:       float = 0.0
    tray_y:       float = 0.0


class Session:
    def __init__(self, session_id: Optional[str] = None):
        self.session_id = session_id or datetime.now().strftime("S%Y%m%d%H%M")
        self.started_at  = datetime.now()
        self.records: list[DiamondRecord] = []
        self._counter = 0

    @property
    def sorted_count(self) -> int:
        return len(self.records)

    @property
    def total_weight_ct(self) -> float:
        return sum(r.weight_ct for r in self.records)

    @property
    def avg_weight_ct(self) -> float:
        return self.total_weight_ct / max(1, self.sorted_count)

    def next_diamond_id(self) -> str:
        self._counter += 1
        return f"{self.session_id}-D{self._counter:04d}"

    def add_record(self, slot_index: int, weight_ct: float,
                   tray_x: float = 0, tray_y: float = 0) -> DiamondRecord:
        rec = DiamondRecord(
            diamond_id  = self.next_diamond_id(),
            slot_index  = slot_index,
            weight_ct   = weight_ct,
            timestamp   = datetime.now().strftime("%H:%M:%S"),
            tray_x      = tray_x,
            tray_y      = tray_y,
        )
        self.records.append(rec)
        return rec

    def export_xlsx(self) -> str:
        os.makedirs(EXPORTS_DIR, exist_ok=True)
        filename = os.path.join(
            EXPORTS_DIR,
            f"{self.session_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Diamond Sort"

        # ── Styles ──────────────────────────────────────────────────────────
        hdr_fill  = PatternFill("solid", fgColor="141714")
        hdr_font  = Font(name="Consolas", bold=True, size=9, color="7a877a")
        cell_font = Font(name="Consolas", size=9, color="d4d9d4")
        grn_font  = Font(name="Consolas", bold=True, size=9, color="3ddc84")
        center    = Alignment(horizontal="center", vertical="center")
        thin      = Side(style="thin", color="1e221e")
        border    = Border(bottom=thin)

        headers = ["#", "Diamond ID", "Weight (ct)", "Slot", "Time", "Tray X", "Tray Y"]
        col_w   = [6, 22, 14, 8, 10, 10, 10]

        # ── Session header block ─────────────────────────────────────────────
        ws.merge_cells("A1:G1")
        ws["A1"] = f"LUMINAX SORTER — {self.session_id}"
        ws["A1"].font = Font(name="Consolas", bold=True, size=11, color="3ddc84")
        ws["A1"].fill = PatternFill("solid", fgColor="080a08")
        ws["A1"].alignment = center

        ws.merge_cells("A2:G2")
        ws["A2"] = (f"Started: {self.started_at.strftime('%Y-%m-%d %H:%M')}  |  "
                    f"Diamonds: {self.sorted_count}  |  "
                    f"Total: {self.total_weight_ct:.4f} ct  |  "
                    f"Avg: {self.avg_weight_ct:.4f} ct")
        ws["A2"].font = Font(name="Consolas", size=9, color="7a877a")
        ws["A2"].fill = PatternFill("solid", fgColor="0e110e")
        ws["A2"].alignment = center
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 18

        # ── Column headers ───────────────────────────────────────────────────
        for ci, (h, w) in enumerate(zip(headers, col_w), start=1):
            cell = ws.cell(row=4, column=ci, value=h)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = center
            cell.border    = border
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[4].height = 18

        # ── Data rows sorted by weight ────────────────────────────────────────
        sorted_records = sorted(self.records, key=lambda r: r.weight_ct)
        for ri, rec in enumerate(sorted_records, start=5):
            row = [
                ri - 4,
                rec.diamond_id,
                round(rec.weight_ct, 4),
                SLOT_ORDER[rec.slot_index] if rec.slot_index < len(SLOT_ORDER) else str(rec.slot_index),
                rec.timestamp,
                round(rec.tray_x, 1),
                round(rec.tray_y, 1),
            ]
            for ci, val in enumerate(row, start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font      = grn_font if ci == 3 else cell_font
                cell.alignment = center
                cell.border    = border
                if ri % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="0e110e")

        # ── Freeze header ────────────────────────────────────────────────────
        ws.freeze_panes = "A5"

        wb.save(filename)
        return filename
